"""Dapodik Ticketing System - FastAPI backend."""
from __future__ import annotations

from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import base64
import logging
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional, Literal

from fastapi import FastAPI, APIRouter, Depends, HTTPException, Query, Response, Request
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from pydantic import BaseModel, Field, EmailStr

from auth import (
    hash_password,
    verify_password,
    create_access_token,
    get_current_user,
    require_role,
)
from fonnte import (
    send_whatsapp,
    send_whatsapp_many,
    fetch_device_info,
    msg_ticket_created,
    msg_status_change,
    msg_new_comment,
    msg_bulk_status,
    normalize_phone,
)

# ---------- Setup ----------
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="Dapodik Ticketing API")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("dapodik")

# ---------- Constants ----------
STATUSES = ["Draft", "Diajukan", "Diproses", "Menunggu Dokumen", "Revisi", "Disetujui", "Selesai", "Ditolak"]
PRIORITIES = ["Rendah", "Normal", "Tinggi", "Mendesak"]
ROLES = ["operator", "koordinator"]

DEFAULT_SERVICES = [
    ("Approval Perubahan Status Kepegawaian Dapodik", 5,
     ["SK Penugasan", "Scan KTP"], False),
    ("Approval Request Hapus Akun PTK", 2,
     [], False),
    ("Approval Permintaan Reset Akun PTK", 1,
     [], False),
    ("Approval Permintaan Reset Akun Sekolah", 1,
     [], False),
    ("Approval Input Siswa Pindah Rombel", 3,
     [], False),
    ("Approval Perubahan Jabatan PTK", 3,
     ["SK Jabatan", "Scan KTP"], False),
    ("Approval Input Siswa Baru", 3,
     [], False),
    ("Approval Penugasan Kepala Sekolah", 5,
     ["SK Penugasan", "Scan KTP"], True),
    ("Approval Pengajuan Mutasi Guru", 3,
     ["SK Penugasan", "Scan KTP"], True),
    ("Approval Input Jam Tambahan di Sekolah Lain", 3,
     ["SK Penugasan", "Scan KTP"], True),
    ("Approval Input Kenaikan Gaji Berkala atau Kenaikan Pangkat", 5,
     ["SK Kenaikan Gaji/Pangkat", "Scan KTP"], True),
]

DEFAULT_KECAMATAN = ["Bogor Tengah", "Bogor Utara", "Bogor Selatan", "Bogor Barat", "Bogor Timur", "Tanah Sareal"]


# Common field templates (reused across schemas)
def F(key, label, type_="text", required=True, options=None, help_text=None, placeholder=None):
    f = {"key": key, "label": label, "type": type_, "required": required, "options": options or [], "help_text": help_text}
    if placeholder:
        f["placeholder"] = placeholder
    return f


COMMON_HEADER = [
    F("nama_sekolah", "Nama Sekolah", "text", help_text="Otomatis terisi dari akun"),
    F("nama_operator", "Nama Operator", "text", help_text="Otomatis terisi dari akun"),
    F("no_whatsapp", "No WhatsApp Aktif Operator", "tel", placeholder="08xxxxxxxxxx"),
]
COMMON_HEADER_WITH_NPSN = [
    F("nama_sekolah", "Nama Sekolah", "text", help_text="Otomatis terisi dari akun"),
    F("npsn", "NPSN", "text", placeholder="20XXXXXX"),
    F("nama_operator", "Nama Operator", "text", help_text="Otomatis terisi dari akun"),
    F("no_whatsapp", "No WhatsApp Aktif Operator", "tel", placeholder="08xxxxxxxxxx"),
]

DEFAULT_FORM_SCHEMAS = {
    "Approval Perubahan Status Kepegawaian Dapodik": COMMON_HEADER + [
        F("nama_ptk", "Nama PTK Sesuai SK", "text"),
        F("nip", "NIP", "text"),
        F("nik", "NIK", "text", placeholder="16 digit NIK"),
        F("no_sk", "No SK Pengangkatan", "text"),
        F("tanggal_sk", "Tanggal Surat (SK)", "date"),
        F("tmt_tugas", "TMT Tugas", "date"),
        F("keterangan_perubahan", "Keterangan Perubahan", "textarea"),
        F("upload_sk_ktp", "Upload SK dan KTP (jadi 1 PDF)", "text", required=False, help_text="Unggah file PDF pada bagian lampiran di bawah"),
    ],
    "Approval Request Hapus Akun PTK": COMMON_HEADER_WITH_NPSN + [
        F("nama_gtk", "Nama GTK", "text"),
        F("nip_gtk", "NIP GTK", "text"),
        F("nik_gtk", "NIK GTK", "text"),
        F("email_akun", "Email atau Akun yang Akan Dihapus", "email"),
    ],
    "Approval Permintaan Reset Akun PTK": COMMON_HEADER_WITH_NPSN + [
        F("nama_gtk", "Nama GTK", "text"),
        F("nip_gtk", "NIP GTK", "text"),
        F("nik_gtk", "NIK GTK", "text"),
        F("email_akun", "Email atau Akun yang Akan Direset", "email"),
    ],
    "Approval Permintaan Reset Akun Sekolah": [
        F("nama_sekolah", "Nama Sekolah", "text", help_text="Otomatis terisi dari akun"),
        F("npsn_npyp", "NPSN / NPYP", "text"),
        F("nama_operator", "Nama Operator", "text", help_text="Otomatis terisi dari akun"),
        F("no_whatsapp", "No WhatsApp Aktif Operator", "tel"),
        F("email_akun", "Email atau Akun yang Akan Direset", "email"),
        F("jenis_akun", "Jenis Akun", "select", options=["Akun Dapodik", "Akun SDM"]),
    ],
    "Approval Input Siswa Pindah Rombel": COMMON_HEADER_WITH_NPSN + [
        F("nama_siswa", "Nama Lengkap Siswa", "text"),
        F("nisn", "NISN", "text"),
        F("nik", "NIK", "text"),
        F("rombel_saat_ini", "Rombel Saat Ini", "text"),
        F("rombel_tujuan", "Rombel Tujuan", "text"),
    ],
    "Approval Perubahan Jabatan PTK": COMMON_HEADER + [
        F("nama_ptk", "Nama PTK Sesuai SK", "text"),
        F("nip", "NIP", "text"),
        F("nik", "NIK", "text"),
        F("no_sk", "No SK Pengangkatan", "text"),
        F("tanggal_sk", "Tanggal Surat (SK)", "date"),
        F("tmt_tugas", "TMT Tugas", "date"),
        F("mata_pelajaran", "Mata Pelajaran", "text"),
        F("upload_sk_ktp", "Upload SK dan KTP (jadi 1 PDF)", "text", required=False, help_text="Unggah file PDF pada bagian lampiran"),
    ],
    "Approval Input Siswa Baru": COMMON_HEADER + [
        F("nama_siswa", "Nama Lengkap Siswa", "text"),
        F("nisn", "NISN", "text"),
        F("nik", "NIK", "text"),
        F("rombel", "Rombel", "text"),
        F("keterangan", "Keterangan", "select", options=["Siswa Baru", "Mutasi"]),
    ],
    "Approval Penugasan Kepala Sekolah": COMMON_HEADER + [
        F("nama_kepsek", "Nama Kepala Sekolah / PLT Sesuai SK", "text"),
        F("nip", "NIP", "text"),
        F("nik", "NIK", "text"),
        F("no_sk", "No SK Pengangkatan", "text"),
        F("tanggal_sk", "Tanggal Surat (SK)", "date"),
        F("tmt_tugas", "TMT Tugas", "date"),
        F("diangkat_sebagai", "Diangkat Sebagai", "select", options=["Kepala Sekolah", "PLT Kepala Sekolah"]),
        F("mapel", "Mapel", "text"),
        F("upload_sk_ktp", "Upload SK dan KTP (jadi 1 PDF)", "text", required=False, help_text="Unggah file PDF pada bagian lampiran"),
    ],
    "Approval Pengajuan Mutasi Guru": COMMON_HEADER + [
        F("nama_gtk", "Nama GTK Sesuai SK", "text"),
        F("nip", "NIP", "text"),
        F("nik", "NIK", "text"),
        F("nuptk", "NUPTK", "text"),
        F("no_sk", "No SK Pengangkatan", "text"),
        F("tanggal_sk", "Tanggal Surat (SK)", "date"),
        F("tmt_tugas", "TMT Tugas", "date"),
        F("diangkat_sebagai", "Diangkat Sebagai", "select", options=["Guru", "Tenaga Kependidikan"]),
        F("mapel", "Mapel", "text"),
        F("unit_kerja_lama", "Unit Kerja Lama", "text"),
        F("unit_kerja_baru", "Unit Kerja Baru", "text"),
        F("upload_sk_ktp", "Upload SK dan KTP (jadi 1 PDF)", "text", required=False, help_text="Unggah file PDF pada bagian lampiran"),
    ],
    "Approval Input Jam Tambahan di Sekolah Lain": COMMON_HEADER + [
        F("nama_gtk", "Nama GTK Sesuai SK", "text"),
        F("nip", "NIP", "text"),
        F("nik", "NIK", "text"),
        F("no_sk", "No SK Pengangkatan", "text"),
        F("tanggal_sk", "Tanggal Surat (SK)", "date"),
        F("tmt_tugas", "TMT Tugas", "date"),
        F("mapel", "Mapel", "text"),
        F("unit_kerja_asal", "Unit Kerja Asal", "text"),
        F("unit_kerja_tujuan", "Unit Kerja Tujuan", "text"),
        F("upload_sk_ktp", "Upload SK dan KTP (jadi 1 PDF)", "text", required=False, help_text="Unggah file PDF pada bagian lampiran"),
    ],
    "Approval Input Kenaikan Gaji Berkala atau Kenaikan Pangkat": COMMON_HEADER + [
        F("nama_gtk", "Nama GTK Sesuai SK", "text"),
        F("nip", "NIP", "text"),
        F("nik", "NIK", "text"),
        F("no_sk_berkala", "No SK Berkala / KP", "text"),
        F("tanggal_sk", "Tanggal Surat (SK)", "date"),
        F("tmt", "TMT", "date"),
        F("gaji_pokok", "Gaji Pokok Sesuai SK", "number"),
        F("upload_sk_ktp", "Upload SK dan KTP (jadi 1 PDF)", "text", required=False, help_text="Unggah file PDF pada bagian lampiran"),
    ],
}


# Keys that must always be optional (not required) — per spec
OPTIONAL_FIELD_KEYS = {"nip", "nik", "nuptk", "nip_gtk", "nik_gtk"}


def _normalize_schema(schema: list) -> list:
    """Mutate schema: ensure NIP/NIK/NUPTK fields are non-required."""
    for f in schema:
        if isinstance(f, dict) and f.get("key") in OPTIONAL_FIELD_KEYS:
            f["required"] = False
    return schema


# Apply normalization to default schemas at import time
for _nama, _sch in DEFAULT_FORM_SCHEMAS.items():
    _normalize_schema(_sch)


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def iso(dt: datetime) -> str:
    return dt.isoformat()


def doc_to_public(doc: dict, drop: tuple = ("password_hash",)) -> dict:
    if not doc:
        return doc
    doc = dict(doc)
    doc["id"] = str(doc.pop("_id"))
    for k in drop:
        doc.pop(k, None)
    return doc


# ---------- Pydantic Schemas ----------
class LoginIn(BaseModel):
    email: EmailStr
    password: str


class ChangePasswordIn(BaseModel):
    old_password: str
    new_password: str = Field(min_length=6)


class KecamatanIn(BaseModel):
    nama: str


class SekolahIn(BaseModel):
    nama: str
    npsn: Optional[str] = None
    kecamatan: str
    jenjang: Optional[str] = None
    alamat: Optional[str] = None


class OperatorIn(BaseModel):
    name: str
    email: EmailStr
    password: str = Field(min_length=6)
    sekolah_id: str
    phone: Optional[str] = None


class OperatorUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    password: Optional[str] = Field(default=None, min_length=6)
    sekolah_id: Optional[str] = None
    phone: Optional[str] = None
    active: Optional[bool] = None


class FormField(BaseModel):
    key: str
    label: str
    type: Literal["text", "number", "date", "select", "textarea", "tel", "email"] = "text"
    required: bool = True
    options: List[str] = []
    help_text: Optional[str] = None
    placeholder: Optional[str] = None


class LayananIn(BaseModel):
    nama: str
    sla_days: int = Field(ge=1, le=60)
    deskripsi: Optional[str] = None
    checklist: List[str] = []
    form_schema: List[FormField] = []
    attachment_required: bool = False


class AttachmentIn(BaseModel):
    filename: str
    mime: str
    data_base64: str  # raw base64 (no data: prefix)


class ChecklistItem(BaseModel):
    label: str
    checked: bool = False


class TicketCreate(BaseModel):
    layanan_id: str
    judul: str
    deskripsi: str
    prioritas: Literal["Rendah", "Normal", "Tinggi", "Mendesak"] = "Normal"
    attachments: List[AttachmentIn] = []
    checklist_state: List[ChecklistItem] = []
    form_data: dict = {}


class AssignIn(BaseModel):
    assignee_id: Optional[str] = None


class ChecklistUpdate(BaseModel):
    items: List[ChecklistItem]


class KbCategoryIn(BaseModel):
    nama: str
    deskripsi: Optional[str] = None


class KbArticleIn(BaseModel):
    title: str
    kategori: Optional[str] = None
    content: str
    tags: List[str] = []


class StatusChange(BaseModel):
    status: Literal["Draft", "Diajukan", "Diproses", "Menunggu Dokumen", "Revisi", "Disetujui", "Selesai", "Ditolak"]
    catatan: Optional[str] = None


class CommentIn(BaseModel):
    content: str


class BulkAssignIn(BaseModel):
    ticket_ids: List[str] = Field(min_length=1, max_length=500)
    assignee_id: Optional[str] = None


class BulkStatusIn(BaseModel):
    ticket_ids: List[str] = Field(min_length=1, max_length=500)
    status: Literal["Draft", "Diajukan", "Diproses", "Menunggu Dokumen", "Revisi", "Disetujui", "Selesai", "Ditolak"]
    catatan: Optional[str] = None


class BulkDeleteIn(BaseModel):
    ticket_ids: List[str] = Field(min_length=1, max_length=500)


class BulkPriorityIn(BaseModel):
    ticket_ids: List[str] = Field(min_length=1, max_length=500)
    prioritas: Literal["Rendah", "Normal", "Tinggi", "Mendesak"]


# ---------- Helpers ----------
async def next_ticket_number() -> str:
    year = datetime.now(timezone.utc).year
    res = await db.counters.find_one_and_update(
        {"_id": f"tickets-{year}"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = (res or {}).get("seq", 1)
    return f"TCK-{year}-{seq:06d}"


async def log_activity(ticket_id: str, actor: dict, kind: str, message: str, meta: Optional[dict] = None):
    await db.activities.insert_one({
        "ticket_id": ticket_id,
        "actor_id": actor["id"],
        "actor_name": actor.get("name", actor.get("email")),
        "actor_role": actor.get("role"),
        "kind": kind,
        "message": message,
        "meta": meta or {},
        "created_at": iso(now_utc()),
    })


async def notify(user_id: str, ticket_id: Optional[str], title: str, body: str):
    await db.notifications.insert_one({
        "user_id": user_id,
        "ticket_id": ticket_id,
        "title": title,
        "body": body,
        "read": False,
        "created_at": iso(now_utc()),
    })


async def log_audit(actor: dict, entity: str, entity_id: Optional[str], action: str, summary: str, meta: Optional[dict] = None):
    """Append-only audit log of admin / system actions."""
    await db.audit_logs.insert_one({
        "actor_id": actor.get("id"),
        "actor_name": actor.get("name", actor.get("email")),
        "actor_role": actor.get("role"),
        "entity": entity,            # e.g. "sekolah", "operator", "layanan", "kecamatan", "ticket", "kb_article"
        "entity_id": entity_id,
        "action": action,            # e.g. "create", "update", "delete", "status_change", "assign"
        "summary": summary,
        "meta": meta or {},
        "created_at": iso(now_utc()),
    })


async def notify_operator_wa(
    ticket: dict,
    event_type: str,
    message: str,
) -> dict:
    """Send a WhatsApp message to the ticket's operator (form_data.no_whatsapp),
    honoring the operator's opt-out preference, and log the result to db.wa_logs.

    Returns the Fonnte response dict (or skip-dict)."""
    phone = (ticket.get("form_data") or {}).get("no_whatsapp")
    operator_id = ticket.get("operator_id")
    log_doc: dict = {
        "ticket_id": str(ticket.get("_id")) if ticket.get("_id") else ticket.get("id"),
        "ticket_number": ticket.get("ticket_number"),
        "operator_id": operator_id,
        "event_type": event_type,
        "phone_last4": (normalize_phone(phone) or "")[-4:] if phone else "",
        "created_at": iso(now_utc()),
    }
    # operator opt-out check
    if operator_id:
        try:
            op = await db.users.find_one({"_id": ObjectId(operator_id)})
            if op and op.get("wa_opt_out"):
                log_doc.update({"status": False, "detail": "operator_opted_out", "skipped": True})
                await db.wa_logs.insert_one(log_doc)
                return {"status": False, "detail": "operator_opted_out", "skipped": True}
        except Exception:
            pass
    if not phone:
        log_doc.update({"status": False, "detail": "no_phone", "skipped": True})
        await db.wa_logs.insert_one(log_doc)
        return {"status": False, "detail": "no_phone", "skipped": True}
    res = await send_whatsapp(phone, message)
    log_doc.update({
        "status": bool(res.get("status")),
        "detail": str(res.get("detail") or "")[:300],
        "fonnte_requestid": res.get("requestid"),
        "quota_remaining": res.get("quota_remaining"),
        "skipped": bool(res.get("skipped")),
    })
    await db.wa_logs.insert_one(log_doc)
    # check quota threshold and alert koordinators if low
    if res.get("quota_remaining") is not None:
        try:
            await maybe_alert_low_quota(int(res["quota_remaining"]))
        except Exception as e:
            logger.warning("low quota alert failed: %s", e)
    return res


QUOTA_LOW_THRESHOLD = 50
QUOTA_ALERT_COOLDOWN_HOURS = 6


async def maybe_alert_low_quota(quota: int):
    """Fire in-app + WhatsApp alert to all koordinator when quota drops below
    threshold. Self-throttled: at most one alert per cooldown window."""
    if quota >= QUOTA_LOW_THRESHOLD:
        return
    state = await db.system_state.find_one({"_id": "quota_alert"})
    now = now_utc()
    if state and state.get("last_sent_at"):
        try:
            last = datetime.fromisoformat(state["last_sent_at"])
            if (now - last).total_seconds() < QUOTA_ALERT_COOLDOWN_HOURS * 3600:
                return
        except Exception:
            pass
    title = "Quota WhatsApp Hampir Habis"
    body = f"Sisa quota Fonnte tinggal {quota} pesan. Segera top-up untuk menjaga kelancaran notifikasi."
    msg = (
        "*Dapodik Ticketing — Peringatan Quota*\n"
        f"Sisa quota WhatsApp (Fonnte): *{quota} pesan*.\n"
        "Mohon segera top-up agar notifikasi ke operator tidak terganggu."
    )
    # in-app to all koordinators + WA to those with phone
    async for k in db.users.find({"role": "koordinator"}):
        kid = str(k["_id"])
        await notify(kid, None, title, body)
        kphone = (k.get("phone") or "").strip()
        if kphone:
            try:
                await send_whatsapp(kphone, msg)
            except Exception as e:
                logger.warning("low-quota WA to koord failed: %s", e)
    await db.system_state.update_one(
        {"_id": "quota_alert"},
        {"$set": {"last_sent_at": iso(now), "quota_at_alert": quota}},
        upsert=True,
    )


async def compute_due_at(layanan_id: str, submitted_at: datetime) -> Optional[str]:
    try:
        lay = await db.services.find_one({"_id": ObjectId(layanan_id)})
    except Exception:
        return None
    if not lay:
        return None
    return iso(submitted_at + timedelta(days=int(lay.get("sla_days", 3))))


def sla_state(due_at_iso: Optional[str], status: str) -> str:
    if status in ("Selesai", "Disetujui", "Ditolak"):
        return "selesai"
    if not due_at_iso:
        return "tidak_diatur"
    due = datetime.fromisoformat(due_at_iso)
    now = now_utc()
    if now > due:
        return "terlambat"
    remaining = (due - now).total_seconds()
    if remaining < 24 * 3600:
        return "hampir_terlambat"
    return "tepat_waktu"


async def enrich_ticket(t: dict) -> dict:
    out = doc_to_public(t)
    out["sla_state"] = sla_state(out.get("due_at"), out.get("status"))
    return out


# ---------- Auth Endpoints ----------
@api.post("/auth/login")
async def login(payload: LoginIn, response: Response):
    user = await db.users.find_one({"email": payload.email.lower()})
    if not user or not user.get("active", True) or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email atau password salah")
    user_id = str(user["_id"])
    token = create_access_token(user_id, user["email"], user["role"])
    response.set_cookie(
        key="access_token", value=token, httponly=True, secure=False, samesite="lax",
        max_age=12 * 3600, path="/",
    )
    return {
        "token": token,
        "user": doc_to_public(user),
    }


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    # enrich with school
    if user.get("sekolah_id"):
        try:
            s = await db.sekolah.find_one({"_id": ObjectId(user["sekolah_id"])})
            if s:
                user["sekolah"] = doc_to_public(s)
        except Exception:
            pass
    return user


@api.post("/auth/change-password")
async def change_password(payload: ChangePasswordIn, user: dict = Depends(get_current_user)):
    doc = await db.users.find_one({"_id": ObjectId(user["id"])})
    if not doc or not verify_password(payload.old_password, doc["password_hash"]):
        raise HTTPException(status_code=400, detail="Password lama tidak sesuai")
    await db.users.update_one(
        {"_id": ObjectId(user["id"])},
        {"$set": {"password_hash": hash_password(payload.new_password), "updated_at": iso(now_utc())}},
    )
    return {"ok": True}


# ---------- Kecamatan ----------
@api.get("/kecamatan")
async def list_kecamatan(_: dict = Depends(get_current_user)):
    items = await db.kecamatan.find().sort("nama", 1).to_list(500)
    return [doc_to_public(i) for i in items]


@api.post("/kecamatan")
async def create_kecamatan(payload: KecamatanIn, user: dict = Depends(require_role("koordinator"))):
    exists = await db.kecamatan.find_one({"nama": payload.nama})
    if exists:
        raise HTTPException(status_code=400, detail="Kecamatan sudah ada")
    res = await db.kecamatan.insert_one({"nama": payload.nama, "created_at": iso(now_utc())})
    await log_audit(user, "kecamatan", str(res.inserted_id), "create", f"Kecamatan '{payload.nama}' ditambahkan")
    return doc_to_public(await db.kecamatan.find_one({"_id": res.inserted_id}))


@api.delete("/kecamatan/{kid}")
async def delete_kecamatan(kid: str, user: dict = Depends(require_role("koordinator"))):
    doc = await db.kecamatan.find_one({"_id": ObjectId(kid)})
    await db.kecamatan.delete_one({"_id": ObjectId(kid)})
    if doc:
        await log_audit(user, "kecamatan", kid, "delete", f"Kecamatan '{doc.get('nama')}' dihapus")
    return {"ok": True}


# ---------- Sekolah ----------
@api.get("/sekolah")
async def list_sekolah(q: Optional[str] = None, kecamatan: Optional[str] = None, _: dict = Depends(get_current_user)):
    filt: dict = {}
    if q:
        filt["nama"] = {"$regex": q, "$options": "i"}
    if kecamatan:
        filt["kecamatan"] = kecamatan
    items = await db.sekolah.find(filt).sort("nama", 1).to_list(2000)
    return [doc_to_public(i) for i in items]


@api.post("/sekolah")
async def create_sekolah(payload: SekolahIn, user: dict = Depends(require_role("koordinator"))):
    doc = payload.model_dump()
    doc["created_at"] = iso(now_utc())
    res = await db.sekolah.insert_one(doc)
    await log_audit(user, "sekolah", str(res.inserted_id), "create", f"Sekolah '{payload.nama}' ditambahkan")
    return doc_to_public(await db.sekolah.find_one({"_id": res.inserted_id}))


@api.put("/sekolah/{sid}")
async def update_sekolah(sid: str, payload: SekolahIn, user: dict = Depends(require_role("koordinator"))):
    await db.sekolah.update_one({"_id": ObjectId(sid)}, {"$set": {**payload.model_dump(), "updated_at": iso(now_utc())}})
    await log_audit(user, "sekolah", sid, "update", f"Sekolah '{payload.nama}' diperbarui")
    return doc_to_public(await db.sekolah.find_one({"_id": ObjectId(sid)}))


@api.delete("/sekolah/{sid}")
async def delete_sekolah(sid: str, user: dict = Depends(require_role("koordinator"))):
    doc = await db.sekolah.find_one({"_id": ObjectId(sid)})
    await db.sekolah.delete_one({"_id": ObjectId(sid)})
    if doc:
        await log_audit(user, "sekolah", sid, "delete", f"Sekolah '{doc.get('nama')}' dihapus")
    return {"ok": True}


# ---------- Operators (users) ----------
@api.get("/operators")
async def list_operators(q: Optional[str] = None, _: dict = Depends(require_role("koordinator"))):
    filt = {"role": "operator"}
    if q:
        filt["$or"] = [{"name": {"$regex": q, "$options": "i"}}, {"email": {"$regex": q, "$options": "i"}}]
    items = await db.users.find(filt).sort("name", 1).to_list(2000)
    # enrich with sekolah
    out = []
    for it in items:
        pub = doc_to_public(it)
        if pub.get("sekolah_id"):
            try:
                s = await db.sekolah.find_one({"_id": ObjectId(pub["sekolah_id"])})
                if s:
                    pub["sekolah"] = doc_to_public(s)
            except Exception:
                pass
        out.append(pub)
    return out


@api.post("/operators")
async def create_operator(payload: OperatorIn, user: dict = Depends(require_role("koordinator"))):
    email_low = payload.email.lower()
    if await db.users.find_one({"email": email_low}):
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    # ensure one operator per sekolah
    if await db.users.find_one({"role": "operator", "sekolah_id": payload.sekolah_id}):
        raise HTTPException(status_code=400, detail="Sekolah ini sudah memiliki akun operator")
    sekolah = await db.sekolah.find_one({"_id": ObjectId(payload.sekolah_id)})
    if not sekolah:
        raise HTTPException(status_code=404, detail="Sekolah tidak ditemukan")
    doc = {
        "name": payload.name,
        "email": email_low,
        "password_hash": hash_password(payload.password),
        "role": "operator",
        "sekolah_id": payload.sekolah_id,
        "phone": payload.phone,
        "active": True,
        "created_at": iso(now_utc()),
    }
    res = await db.users.insert_one(doc)
    await log_audit(user, "operator", str(res.inserted_id), "create", f"Operator '{payload.name}' ({email_low}) ditambahkan")
    return doc_to_public(await db.users.find_one({"_id": res.inserted_id}))


@api.put("/operators/{uid}")
async def update_operator(uid: str, payload: OperatorUpdate, user: dict = Depends(require_role("koordinator"))):
    update: dict = {}
    data = payload.model_dump(exclude_unset=True)
    if "password" in data and data["password"]:
        update["password_hash"] = hash_password(data.pop("password"))
    if "email" in data and data["email"]:
        data["email"] = data["email"].lower()
    update.update(data)
    update["updated_at"] = iso(now_utc())
    if "sekolah_id" in update and update["sekolah_id"]:
        clash = await db.users.find_one({
            "role": "operator",
            "sekolah_id": update["sekolah_id"],
            "_id": {"$ne": ObjectId(uid)},
        })
        if clash:
            raise HTTPException(status_code=400, detail="Sekolah ini sudah memiliki akun operator")
    if "email" in update and update["email"]:
        clash = await db.users.find_one({
            "email": update["email"],
            "_id": {"$ne": ObjectId(uid)},
        })
        if clash:
            raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    await db.users.update_one({"_id": ObjectId(uid)}, {"$set": update})
    after = await db.users.find_one({"_id": ObjectId(uid)})
    await log_audit(user, "operator", uid, "update", f"Operator '{(after or {}).get('name')}' diperbarui")
    return doc_to_public(after)


@api.delete("/operators/{uid}")
async def delete_operator(uid: str, user: dict = Depends(require_role("koordinator"))):
    doc = await db.users.find_one({"_id": ObjectId(uid), "role": "operator"})
    await db.users.delete_one({"_id": ObjectId(uid), "role": "operator"})
    if doc:
        await log_audit(user, "operator", uid, "delete", f"Operator '{doc.get('name')}' dihapus")
    return {"ok": True}


# ---------- Layanan ----------
@api.get("/layanan")
async def list_layanan(_: dict = Depends(get_current_user)):
    items = await db.services.find().sort("nama", 1).to_list(500)
    return [doc_to_public(i) for i in items]


@api.post("/layanan")
async def create_layanan(payload: LayananIn, _: dict = Depends(require_role("koordinator"))):
    res = await db.services.insert_one({**payload.model_dump(), "created_at": iso(now_utc())})
    return doc_to_public(await db.services.find_one({"_id": res.inserted_id}))


@api.put("/layanan/{lid}")
async def update_layanan(lid: str, payload: LayananIn, _: dict = Depends(require_role("koordinator"))):
    await db.services.update_one({"_id": ObjectId(lid)}, {"$set": {**payload.model_dump(), "updated_at": iso(now_utc())}})
    return doc_to_public(await db.services.find_one({"_id": ObjectId(lid)}))


@api.delete("/layanan/{lid}")
async def delete_layanan(lid: str, _: dict = Depends(require_role("koordinator"))):
    await db.services.delete_one({"_id": ObjectId(lid)})
    return {"ok": True}


# ---------- Tickets ----------
async def _enrich_one(t: dict) -> dict:
    out = doc_to_public(t)
    out["sla_state"] = sla_state(out.get("due_at"), out.get("status"))
    return out


@api.post("/tickets")
async def create_ticket(payload: TicketCreate, user: dict = Depends(get_current_user)):
    if user["role"] != "operator":
        raise HTTPException(status_code=403, detail="Hanya operator yang dapat membuat pengajuan")
    layanan = await db.services.find_one({"_id": ObjectId(payload.layanan_id)})
    if not layanan:
        raise HTTPException(status_code=404, detail="Jenis layanan tidak ditemukan")
    if layanan.get("attachment_required") and not payload.attachments:
        raise HTTPException(status_code=400, detail="Layanan ini wajib menyertakan lampiran dokumen (SK/KTP).")
    sekolah = None
    if user.get("sekolah_id"):
        sekolah = await db.sekolah.find_one({"_id": ObjectId(user["sekolah_id"])})

    submitted_at = now_utc()
    ticket_number = await next_ticket_number()
    due_at = iso(submitted_at + timedelta(days=int(layanan.get("sla_days", 3))))

    # Build checklist state: use submitted state if present, else seed from layanan.checklist
    checklist_items: List[dict] = []
    if payload.checklist_state:
        checklist_items = [{"label": c.label, "checked": c.checked} for c in payload.checklist_state]
    elif isinstance(layanan.get("checklist"), list):
        checklist_items = [{"label": label, "checked": False} for label in layanan["checklist"]]

    doc = {
        "ticket_number": ticket_number,
        "judul": payload.judul,
        "deskripsi": payload.deskripsi,
        "layanan_id": payload.layanan_id,
        "layanan_nama": layanan["nama"],
        "sla_days": int(layanan.get("sla_days", 3)),
        "prioritas": payload.prioritas,
        "status": "Diajukan",
        "operator_id": user["id"],
        "operator_name": user["name"],
        "sekolah_id": user.get("sekolah_id"),
        "sekolah_nama": (sekolah or {}).get("nama"),
        "kecamatan": (sekolah or {}).get("kecamatan"),
        "submitted_at": iso(submitted_at),
        "due_at": due_at,
        "created_at": iso(submitted_at),
        "updated_at": iso(submitted_at),
        "closed_at": None,
        "assignee_id": None,
        "assignee_name": None,
        "checklist": checklist_items,
        "form_data": payload.form_data or {},
    }
    res = await db.tickets.insert_one(doc)
    tid = str(res.inserted_id)

    # Save attachments
    for att in payload.attachments[:10]:
        await db.attachments.insert_one({
            "ticket_id": tid,
            "filename": att.filename,
            "mime": att.mime,
            "data_base64": att.data_base64,
            "uploaded_by": user["id"],
            "uploaded_at": iso(now_utc()),
        })

    await log_activity(tid, user, "created", f"Pengajuan dibuat: {payload.judul}")
    # notify all koordinator
    async for k in db.users.find({"role": "koordinator"}):
        await notify(str(k["_id"]), tid, "Pengajuan Baru", f"{ticket_number} - {payload.judul}")

    saved = await db.tickets.find_one({"_id": res.inserted_id})
    enriched = await _enrich_one(saved)

    # WhatsApp confirmation to operator (skips silently if opted out or no phone)
    try:
        await notify_operator_wa(saved, "created", msg_ticket_created(saved))
    except Exception as e:
        logger.warning("WA notify (created) failed: %s", e)

    return enriched


@api.get("/tickets")
async def list_tickets(
    q: Optional[str] = None,
    status: Optional[str] = None,
    layanan_id: Optional[str] = None,
    kecamatan: Optional[str] = None,
    sekolah_id: Optional[str] = None,
    operator_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    sla_filter: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
    user: dict = Depends(get_current_user),
):
    filt: dict = {}
    if user["role"] == "operator":
        filt["operator_id"] = user["id"]
    if q:
        filt["$or"] = [
            {"ticket_number": {"$regex": q, "$options": "i"}},
            {"judul": {"$regex": q, "$options": "i"}},
            {"sekolah_nama": {"$regex": q, "$options": "i"}},
            {"operator_name": {"$regex": q, "$options": "i"}},
            {"layanan_nama": {"$regex": q, "$options": "i"}},
        ]
    if status:
        filt["status"] = status
    if layanan_id:
        filt["layanan_id"] = layanan_id
    if kecamatan:
        filt["kecamatan"] = kecamatan
    if sekolah_id:
        filt["sekolah_id"] = sekolah_id
    if operator_id:
        filt["operator_id"] = operator_id
    if from_date or to_date:
        rng = {}
        if from_date:
            rng["$gte"] = from_date
        if to_date:
            rng["$lte"] = to_date + "T23:59:59"
        filt["submitted_at"] = rng

    total = await db.tickets.count_documents(filt)
    cursor = db.tickets.find(filt).sort("submitted_at", -1).skip(skip).limit(limit)
    items = [await _enrich_one(t) async for t in cursor]
    if sla_filter:
        items = [i for i in items if i["sla_state"] == sla_filter]
    return {"items": items, "total": total}


@api.get("/tickets/{tid}")
async def get_ticket(tid: str, user: dict = Depends(get_current_user)):
    t = await db.tickets.find_one({"_id": ObjectId(tid)})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket tidak ditemukan")
    if user["role"] == "operator" and t.get("operator_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Akses ditolak")
    enriched = await _enrich_one(t)
    activities = await db.activities.find({"ticket_id": tid}).sort("created_at", 1).to_list(1000)
    attachments = await db.attachments.find({"ticket_id": tid}, {"data_base64": 0}).sort("uploaded_at", 1).to_list(100)
    enriched["activities"] = [doc_to_public(a) for a in activities]
    enriched["attachments"] = [doc_to_public(a) for a in attachments]
    return enriched


@api.post("/tickets/{tid}/status")
async def change_status(tid: str, payload: StatusChange, user: dict = Depends(require_role("koordinator"))):
    t = await db.tickets.find_one({"_id": ObjectId(tid)})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket tidak ditemukan")
    old = t.get("status")
    update = {"status": payload.status, "updated_at": iso(now_utc())}
    if payload.status in ("Selesai", "Disetujui", "Ditolak"):
        update["closed_at"] = iso(now_utc())
    await db.tickets.update_one({"_id": ObjectId(tid)}, {"$set": update})
    msg = f"Status diubah: {old} → {payload.status}"
    if payload.catatan:
        msg += f" — {payload.catatan}"
    await log_activity(tid, user, "status_change", msg, {"from": old, "to": payload.status, "catatan": payload.catatan})
    await log_audit(user, "ticket", tid, "status_change", f"{t['ticket_number']}: {old} → {payload.status}", {"catatan": payload.catatan})
    await notify(t["operator_id"], tid, "Status Pengajuan Diperbarui", f"{t['ticket_number']}: {payload.status}")
    saved = await db.tickets.find_one({"_id": ObjectId(tid)})

    # WhatsApp notify to operator
    try:
        await notify_operator_wa(saved, "status_change", msg_status_change(saved, old, payload.status, payload.catatan))
    except Exception as e:
        logger.warning("WA notify (status) failed: %s", e)

    return await _enrich_one(saved)


@api.post("/tickets/{tid}/comments")
async def add_comment(tid: str, payload: CommentIn, user: dict = Depends(get_current_user)):
    t = await db.tickets.find_one({"_id": ObjectId(tid)})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket tidak ditemukan")
    if user["role"] == "operator" and t.get("operator_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Akses ditolak")
    await log_activity(tid, user, "comment", payload.content)
    # notify counterpart
    if user["role"] == "operator":
        async for k in db.users.find({"role": "koordinator"}):
            await notify(str(k["_id"]), tid, "Komentar baru dari operator", f"{t['ticket_number']}")
    else:
        await notify(t["operator_id"], tid, "Komentar baru dari koordinator", f"{t['ticket_number']}")
        # WhatsApp: koordinator → operator
        try:
            await notify_operator_wa(t, "comment", msg_new_comment(t, user.get("name") or "Koordinator", payload.content))
        except Exception as e:
            logger.warning("WA notify (comment) failed: %s", e)
    return {"ok": True}


@api.post("/tickets/{tid}/attachments")
async def add_attachment(tid: str, payload: AttachmentIn, user: dict = Depends(get_current_user)):
    t = await db.tickets.find_one({"_id": ObjectId(tid)})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket tidak ditemukan")
    if user["role"] == "operator" and t.get("operator_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Akses ditolak")
    if payload.mime not in ("application/pdf", "image/jpeg", "image/png", "image/jpg"):
        raise HTTPException(status_code=400, detail="Format file tidak didukung")
    # ~5MB raw limit
    if len(payload.data_base64) > 7_500_000:
        raise HTTPException(status_code=400, detail="Ukuran file maksimum 5MB")
    res = await db.attachments.insert_one({
        "ticket_id": tid,
        "filename": payload.filename,
        "mime": payload.mime,
        "data_base64": payload.data_base64,
        "uploaded_by": user["id"],
        "uploaded_at": iso(now_utc()),
    })
    await log_activity(tid, user, "attachment", f"Mengunggah dokumen: {payload.filename}")
    a = await db.attachments.find_one({"_id": res.inserted_id}, {"data_base64": 0})
    return doc_to_public(a)


@api.get("/tickets/{tid}/attachments/{aid}/download")
async def download_attachment(tid: str, aid: str, user: dict = Depends(get_current_user)):
    t = await db.tickets.find_one({"_id": ObjectId(tid)})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket tidak ditemukan")
    if user["role"] == "operator" and t.get("operator_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Akses ditolak")
    a = await db.attachments.find_one({"_id": ObjectId(aid), "ticket_id": tid})
    if not a:
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    data = base64.b64decode(a["data_base64"])
    return StreamingResponse(
        io.BytesIO(data),
        media_type=a["mime"],
        headers={"Content-Disposition": f'attachment; filename="{a["filename"]}"'},
    )


# ---------- Notifications ----------
@api.get("/notifications")
async def list_notifications(user: dict = Depends(get_current_user)):
    items = await db.notifications.find({"user_id": user["id"]}).sort("created_at", -1).limit(50).to_list(50)
    unread = await db.notifications.count_documents({"user_id": user["id"], "read": False})
    return {"items": [doc_to_public(i) for i in items], "unread": unread}


@api.post("/notifications/mark-read")
async def mark_read(user: dict = Depends(get_current_user)):
    await db.notifications.update_many({"user_id": user["id"], "read": False}, {"$set": {"read": True}})
    return {"ok": True}


# ---------- Ticket Assignment & Checklist ----------
@api.post("/tickets/{tid}/assign")
async def assign_ticket(tid: str, payload: AssignIn, user: dict = Depends(require_role("koordinator"))):
    t = await db.tickets.find_one({"_id": ObjectId(tid)})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket tidak ditemukan")
    assignee_name = None
    if payload.assignee_id:
        a = await db.users.find_one({"_id": ObjectId(payload.assignee_id), "role": "koordinator"})
        if not a:
            raise HTTPException(status_code=404, detail="Petugas tidak ditemukan")
        assignee_name = a.get("name") or a.get("email")
    await db.tickets.update_one(
        {"_id": ObjectId(tid)},
        {"$set": {"assignee_id": payload.assignee_id, "assignee_name": assignee_name, "updated_at": iso(now_utc())}},
    )
    msg = f"Tugas dialihkan ke {assignee_name}" if assignee_name else "Penugasan dilepas"
    await log_activity(tid, user, "assign", msg, {"assignee_id": payload.assignee_id, "assignee_name": assignee_name})
    await log_audit(user, "ticket", tid, "assign", f"{t['ticket_number']}: {msg}")
    if payload.assignee_id and payload.assignee_id != user["id"]:
        await notify(payload.assignee_id, tid, "Ticket Ditugaskan", f"{t['ticket_number']} - {t['judul']}")
    saved = await db.tickets.find_one({"_id": ObjectId(tid)})
    return await _enrich_one(saved)


@api.post("/tickets/{tid}/checklist")
async def update_checklist(tid: str, payload: ChecklistUpdate, user: dict = Depends(get_current_user)):
    t = await db.tickets.find_one({"_id": ObjectId(tid)})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket tidak ditemukan")
    if user["role"] == "operator" and t.get("operator_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Akses ditolak")
    items = [{"label": c.label, "checked": c.checked} for c in payload.items]
    await db.tickets.update_one(
        {"_id": ObjectId(tid)},
        {"$set": {"checklist": items, "updated_at": iso(now_utc())}},
    )
    await log_activity(tid, user, "checklist", f"Checklist diperbarui ({sum(1 for i in items if i['checked'])}/{len(items)} selesai)")
    saved = await db.tickets.find_one({"_id": ObjectId(tid)})
    return await _enrich_one(saved)


@api.post("/tickets/bulk-assign")
async def bulk_assign(payload: BulkAssignIn, user: dict = Depends(require_role("koordinator"))):
    assignee_name = None
    if payload.assignee_id:
        a = await db.users.find_one({"_id": ObjectId(payload.assignee_id), "role": "koordinator"})
        if not a:
            raise HTTPException(status_code=404, detail="Petugas tidak ditemukan")
        assignee_name = a.get("name") or a.get("email")

    updated = 0
    msg_log = f"Tugas dialihkan ke {assignee_name}" if assignee_name else "Penugasan dilepas"
    for tid in payload.ticket_ids:
        try:
            oid = ObjectId(tid)
        except Exception:
            continue
        t = await db.tickets.find_one({"_id": oid})
        if not t:
            continue
        await db.tickets.update_one(
            {"_id": oid},
            {"$set": {"assignee_id": payload.assignee_id, "assignee_name": assignee_name, "updated_at": iso(now_utc())}},
        )
        await log_activity(tid, user, "assign", msg_log, {"assignee_id": payload.assignee_id, "assignee_name": assignee_name, "bulk": True})
        if payload.assignee_id and payload.assignee_id != user["id"]:
            await notify(payload.assignee_id, tid, "Ticket Ditugaskan", f"{t['ticket_number']} - {t['judul']}")
        updated += 1

    await log_audit(
        user, "ticket", None, "bulk_assign",
        f"Bulk assign {updated} tiket → {assignee_name or 'Tidak ditugaskan'}",
        {"count": updated, "assignee_id": payload.assignee_id, "ticket_ids": payload.ticket_ids[:50]},
    )
    return {"ok": True, "updated": updated}


@api.post("/tickets/bulk-status")
async def bulk_status(payload: BulkStatusIn, user: dict = Depends(require_role("koordinator"))):
    updated = 0
    wa_pending: list[dict] = []
    for tid in payload.ticket_ids:
        try:
            oid = ObjectId(tid)
        except Exception:
            continue
        t = await db.tickets.find_one({"_id": oid})
        if not t:
            continue
        old = t.get("status")
        if old == payload.status:
            continue
        update = {"status": payload.status, "updated_at": iso(now_utc())}
        if payload.status in ("Selesai", "Disetujui", "Ditolak"):
            update["closed_at"] = iso(now_utc())
        await db.tickets.update_one({"_id": oid}, {"$set": update})
        msg = f"Status diubah: {old} → {payload.status}"
        if payload.catatan:
            msg += f" — {payload.catatan}"
        await log_activity(tid, user, "status_change", msg, {"from": old, "to": payload.status, "catatan": payload.catatan, "bulk": True})
        await notify(t["operator_id"], tid, "Status Pengajuan Diperbarui", f"{t['ticket_number']}: {payload.status}")
        wa_pending.append({**t, "status": payload.status})
        updated += 1

    wa_sent = 0
    for t in wa_pending:
        try:
            res = await notify_operator_wa(t, "bulk_status", msg_bulk_status(t, payload.status))
            if res.get("status"):
                wa_sent += 1
        except Exception as e:
            logger.warning("WA bulk-status notify failed: %s", e)

    await log_audit(
        user, "ticket", None, "bulk_status",
        f"Bulk status change {updated} tiket → {payload.status}",
        {"count": updated, "status": payload.status, "catatan": payload.catatan, "ticket_ids": payload.ticket_ids[:50], "wa_sent": wa_sent},
    )
    return {"ok": True, "updated": updated, "wa_sent": wa_sent}


@api.post("/tickets/bulk-delete")
async def bulk_delete(payload: BulkDeleteIn, user: dict = Depends(require_role("koordinator"))):
    deleted = 0
    ticket_numbers = []
    for tid in payload.ticket_ids:
        try:
            oid = ObjectId(tid)
        except Exception:
            continue
        t = await db.tickets.find_one({"_id": oid})
        if not t:
            continue
        ticket_numbers.append(t.get("ticket_number"))
        # cascade-delete related data
        await db.tickets.delete_one({"_id": oid})
        await db.activities.delete_many({"ticket_id": tid})
        await db.attachments.delete_many({"ticket_id": tid})
        await db.notifications.delete_many({"ticket_id": tid})
        deleted += 1

    await log_audit(
        user, "ticket", None, "bulk_delete",
        f"Bulk delete {deleted} tiket: {', '.join(ticket_numbers[:5])}{'...' if len(ticket_numbers) > 5 else ''}",
        {"count": deleted, "ticket_numbers": ticket_numbers[:50]},
    )
    return {"ok": True, "deleted": deleted}


@api.post("/tickets/bulk-priority")
async def bulk_priority(payload: BulkPriorityIn, user: dict = Depends(require_role("koordinator"))):
    updated = 0
    for tid in payload.ticket_ids:
        try:
            oid = ObjectId(tid)
        except Exception:
            continue
        t = await db.tickets.find_one({"_id": oid})
        if not t:
            continue
        old = t.get("prioritas")
        if old == payload.prioritas:
            continue
        await db.tickets.update_one(
            {"_id": oid},
            {"$set": {"prioritas": payload.prioritas, "updated_at": iso(now_utc())}},
        )
        await log_activity(tid, user, "priority_change", f"Prioritas diubah: {old} → {payload.prioritas}", {"from": old, "to": payload.prioritas, "bulk": True})
        updated += 1

    await log_audit(
        user, "ticket", None, "bulk_priority",
        f"Bulk priority {updated} tiket → {payload.prioritas}",
        {"count": updated, "prioritas": payload.prioritas, "ticket_ids": payload.ticket_ids[:50]},
    )
    return {"ok": True, "updated": updated}


# ---------- Knowledge Base ----------
@api.get("/kb/categories")
async def kb_list_categories(_: dict = Depends(get_current_user)):
    items = await db.kb_categories.find().sort("nama", 1).to_list(200)
    return [doc_to_public(i) for i in items]


@api.post("/kb/categories")
async def kb_create_category(payload: KbCategoryIn, user: dict = Depends(require_role("koordinator"))):
    if await db.kb_categories.find_one({"nama": payload.nama}):
        raise HTTPException(status_code=400, detail="Kategori sudah ada")
    res = await db.kb_categories.insert_one({**payload.model_dump(), "created_at": iso(now_utc())})
    await log_audit(user, "kb_category", str(res.inserted_id), "create", f"Kategori KB '{payload.nama}' ditambahkan")
    return doc_to_public(await db.kb_categories.find_one({"_id": res.inserted_id}))


@api.delete("/kb/categories/{cid}")
async def kb_delete_category(cid: str, user: dict = Depends(require_role("koordinator"))):
    doc = await db.kb_categories.find_one({"_id": ObjectId(cid)})
    await db.kb_categories.delete_one({"_id": ObjectId(cid)})
    if doc:
        await log_audit(user, "kb_category", cid, "delete", f"Kategori KB '{doc.get('nama')}' dihapus")
    return {"ok": True}


@api.get("/kb/articles")
async def kb_list_articles(
    q: Optional[str] = None,
    kategori: Optional[str] = None,
    _: dict = Depends(get_current_user),
):
    filt: dict = {}
    if q:
        filt["$or"] = [
            {"title": {"$regex": q, "$options": "i"}},
            {"content": {"$regex": q, "$options": "i"}},
            {"tags": {"$regex": q, "$options": "i"}},
        ]
    if kategori:
        filt["kategori"] = kategori
    items = await db.kb_articles.find(filt, {"content": 0}).sort("updated_at", -1).limit(200).to_list(200)
    return [doc_to_public(i) for i in items]


@api.get("/kb/articles/{aid}")
async def kb_get_article(aid: str, user: dict = Depends(get_current_user)):
    doc = await db.kb_articles.find_one({"_id": ObjectId(aid)})
    if not doc:
        raise HTTPException(status_code=404, detail="Artikel tidak ditemukan")
    # increment view count (non-blocking)
    await db.kb_articles.update_one({"_id": ObjectId(aid)}, {"$inc": {"views": 1}})
    doc["views"] = (doc.get("views") or 0) + 1
    return doc_to_public(doc)


@api.post("/kb/articles")
async def kb_create_article(payload: KbArticleIn, user: dict = Depends(require_role("koordinator"))):
    doc = {
        **payload.model_dump(),
        "author_id": user["id"],
        "author_name": user.get("name"),
        "views": 0,
        "created_at": iso(now_utc()),
        "updated_at": iso(now_utc()),
    }
    res = await db.kb_articles.insert_one(doc)
    await log_audit(user, "kb_article", str(res.inserted_id), "create", f"Artikel KB '{payload.title}' ditambahkan")
    return doc_to_public(await db.kb_articles.find_one({"_id": res.inserted_id}))


@api.put("/kb/articles/{aid}")
async def kb_update_article(aid: str, payload: KbArticleIn, user: dict = Depends(require_role("koordinator"))):
    await db.kb_articles.update_one(
        {"_id": ObjectId(aid)},
        {"$set": {**payload.model_dump(), "updated_at": iso(now_utc())}},
    )
    await log_audit(user, "kb_article", aid, "update", f"Artikel KB '{payload.title}' diperbarui")
    return doc_to_public(await db.kb_articles.find_one({"_id": ObjectId(aid)}))


@api.delete("/kb/articles/{aid}")
async def kb_delete_article(aid: str, user: dict = Depends(require_role("koordinator"))):
    doc = await db.kb_articles.find_one({"_id": ObjectId(aid)})
    await db.kb_articles.delete_one({"_id": ObjectId(aid)})
    if doc:
        await log_audit(user, "kb_article", aid, "delete", f"Artikel KB '{doc.get('title')}' dihapus")
    return {"ok": True}


# ---------- Audit Log ----------
@api.get("/audit")
async def list_audit(
    q: Optional[str] = None,
    entity: Optional[str] = None,
    action: Optional[str] = None,
    actor_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
    _: dict = Depends(require_role("koordinator")),
):
    filt: dict = {}
    if entity:
        filt["entity"] = entity
    if action:
        filt["action"] = action
    if actor_id:
        filt["actor_id"] = actor_id
    if q:
        filt["$or"] = [
            {"summary": {"$regex": q, "$options": "i"}},
            {"actor_name": {"$regex": q, "$options": "i"}},
        ]
    if from_date or to_date:
        rng: dict = {}
        if from_date:
            rng["$gte"] = from_date
        if to_date:
            rng["$lte"] = to_date + "T23:59:59"
        filt["created_at"] = rng
    total = await db.audit_logs.count_documents(filt)
    items = await db.audit_logs.find(filt).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"items": [doc_to_public(i) for i in items], "total": total}


# ---------- Executive (Dashboard Pimpinan) ----------
@api.get("/executive/stats")
async def executive_stats(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    _: dict = Depends(require_role("koordinator")),
):
    base: dict = {}
    if from_date or to_date:
        rng: dict = {}
        if from_date:
            rng["$gte"] = from_date
        if to_date:
            rng["$lte"] = to_date + "T23:59:59"
        base["submitted_at"] = rng

    total = await db.tickets.count_documents(base)
    selesai = await db.tickets.count_documents({**base, "status": {"$in": ["Selesai", "Disetujui"]}})
    ditolak = await db.tickets.count_documents({**base, "status": "Ditolak"})
    diproses = await db.tickets.count_documents({**base, "status": {"$in": ["Diajukan", "Diproses", "Menunggu Dokumen", "Revisi"]}})

    # by kecamatan
    pipeline_kec = [
        {"$match": base},
        {"$group": {"_id": "$kecamatan", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    by_kecamatan = []
    async for row in db.tickets.aggregate(pipeline_kec):
        by_kecamatan.append({"kecamatan": row["_id"] or "—", "count": row["count"]})

    # by layanan
    pipeline_lay = [
        {"$match": base},
        {"$group": {"_id": "$layanan_nama", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    by_layanan = []
    async for row in db.tickets.aggregate(pipeline_lay):
        by_layanan.append({"layanan": row["_id"] or "—", "count": row["count"]})

    # top sekolah
    pipeline_sek = [
        {"$match": base},
        {"$group": {"_id": "$sekolah_nama", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10},
    ]
    top_sekolah = []
    async for row in db.tickets.aggregate(pipeline_sek):
        top_sekolah.append({"sekolah": row["_id"] or "—", "count": row["count"]})

    # avg processing time (hours) for closed tickets
    closed = await db.tickets.find(
        {**base, "closed_at": {"$ne": None}, "submitted_at": {"$ne": None}},
        {"submitted_at": 1, "closed_at": 1, "layanan_nama": 1, "sla_days": 1},
    ).to_list(5000)
    total_h = 0.0
    n = 0
    sla_met = 0
    per_layanan_durations: dict = {}
    for t in closed:
        try:
            s = datetime.fromisoformat(t["submitted_at"])
            c = datetime.fromisoformat(t["closed_at"])
            hours = (c - s).total_seconds() / 3600
            total_h += hours
            n += 1
            sla_h = float(t.get("sla_days") or 3) * 24
            if hours <= sla_h:
                sla_met += 1
            key = t.get("layanan_nama") or "—"
            per_layanan_durations.setdefault(key, []).append(hours)
        except Exception:
            continue
    avg_hours = round(total_h / n, 1) if n else 0
    sla_compliance = round((sla_met / n) * 100, 1) if n else 0
    avg_per_layanan = [
        {"layanan": k, "avg_hours": round(sum(v) / len(v), 1), "count": len(v)}
        for k, v in per_layanan_durations.items()
    ]
    avg_per_layanan.sort(key=lambda x: x["avg_hours"], reverse=True)

    # workload per koordinator (assigned)
    pipeline_assignee = [
        {"$match": {**base, "assignee_id": {"$ne": None}}},
        {"$group": {"_id": {"id": "$assignee_id", "name": "$assignee_name"}, "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    workload = []
    async for row in db.tickets.aggregate(pipeline_assignee):
        workload.append({"assignee_id": row["_id"]["id"], "assignee_name": row["_id"]["name"], "count": row["count"]})

    return {
        "total": total,
        "selesai": selesai,
        "ditolak": ditolak,
        "diproses": diproses,
        "by_kecamatan": by_kecamatan,
        "by_layanan": by_layanan,
        "top_sekolah": top_sekolah,
        "avg_processing_hours": avg_hours,
        "sla_compliance_pct": sla_compliance,
        "avg_per_layanan": avg_per_layanan[:8],
        "workload": workload,
    }


# ---------- Koordinator list (for assignment) ----------
@api.get("/koordinators")
async def list_koordinators(_: dict = Depends(require_role("koordinator"))):
    items = await db.users.find({"role": "koordinator"}).sort("name", 1).to_list(200)
    return [doc_to_public(i) for i in items]


# ---------- Admin Maintenance ----------
class WhatsAppTestIn(BaseModel):
    target: str
    message: Optional[str] = None


class WaPreferenceIn(BaseModel):
    wa_opt_out: Optional[bool] = None
    phone: Optional[str] = None  # WhatsApp number for koordinator self-alerts


@api.patch("/auth/preferences")
async def update_preferences(payload: WaPreferenceIn, user: dict = Depends(get_current_user)):
    update: dict = {"updated_at": iso(now_utc())}
    data = payload.model_dump(exclude_unset=True)
    if "wa_opt_out" in data:
        update["wa_opt_out"] = bool(data["wa_opt_out"])
    if "phone" in data:
        update["phone"] = data["phone"]
    await db.users.update_one({"_id": ObjectId(user["id"])}, {"$set": update})
    fresh = await db.users.find_one({"_id": ObjectId(user["id"])})
    return {"ok": True, "wa_opt_out": bool(fresh.get("wa_opt_out")), "phone": fresh.get("phone")}


@api.post("/admin/test-whatsapp")
async def admin_test_wa(payload: WhatsAppTestIn, user: dict = Depends(require_role("koordinator"))):
    """Send a test WhatsApp message to verify Fonnte connectivity."""
    msg = payload.message or "Tes notifikasi Dapodik Ticketing. Jika Anda menerima ini, integrasi WhatsApp aktif."
    result = await send_whatsapp(payload.target, msg)
    # Log to wa_logs so it appears in delivery stats
    await db.wa_logs.insert_one({
        "ticket_id": None,
        "ticket_number": None,
        "operator_id": None,
        "event_type": "test",
        "phone_last4": (normalize_phone(payload.target) or "")[-4:],
        "status": bool(result.get("status")),
        "detail": str(result.get("detail") or "")[:300],
        "fonnte_requestid": result.get("requestid"),
        "quota_remaining": result.get("quota_remaining"),
        "skipped": bool(result.get("skipped")),
        "created_at": iso(now_utc()),
    })
    await log_audit(user, "system", None, "test_whatsapp", f"Test WA → {payload.target[-4:]}", {"result_status": result.get("status")})
    return result


@api.post("/admin/wa-logs/{log_id}/resend")
async def resend_wa_log(log_id: str, user: dict = Depends(require_role("koordinator"))):
    """Resend a previously-failed WhatsApp message. Rebuilds the message from
    the original event_type using the latest ticket state."""
    try:
        log = await db.wa_logs.find_one({"_id": ObjectId(log_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Log tidak ditemukan")
    if not log:
        raise HTTPException(status_code=404, detail="Log tidak ditemukan")
    event = log.get("event_type") or "manual_resend"

    ticket = None
    if log.get("ticket_id"):
        try:
            ticket = await db.tickets.find_one({"_id": ObjectId(log["ticket_id"])})
        except Exception:
            ticket = None

    if event == "test":
        target_raw = None
        if log.get("phone_last4"):
            target_raw = None  # we cannot reconstruct full phone from last4 — fall back to error
        # cannot reconstruct full phone for plain "test" — refuse
        raise HTTPException(status_code=400, detail="Test message tidak dapat di-resend (nomor lengkap tidak disimpan untuk privasi)")

    if not ticket:
        raise HTTPException(status_code=400, detail="Tiket terkait tidak ditemukan — tidak bisa rebuild pesan")

    phone = (ticket.get("form_data") or {}).get("no_whatsapp")
    if not phone:
        raise HTTPException(status_code=400, detail="Nomor WhatsApp tidak tersedia pada tiket")

    if event == "created":
        message = msg_ticket_created(ticket)
    elif event in ("status_change", "bulk_status"):
        message = msg_bulk_status(ticket, ticket.get("status") or "")
    elif event == "comment":
        message = (
            "*Dapodik Ticketing — Pemberitahuan*\n"
            f"No. Ticket: *{ticket.get('ticket_number')}*\n"
            "Mohon cek aplikasi untuk komentar terbaru dari koordinator."
        )
    else:
        message = (
            "*Dapodik Ticketing — Pemberitahuan*\n"
            f"No. Ticket: *{ticket.get('ticket_number')}*\n"
            f"Status saat ini: *{ticket.get('status')}*"
        )

    res = await notify_operator_wa(ticket, f"resend_{event}", message)
    await log_audit(
        user, "system", str(log["_id"]), "wa_resend",
        f"Resend WA log ({event}) untuk {ticket.get('ticket_number')}",
        {"original_log_id": str(log["_id"]), "result_status": res.get("status")},
    )
    return res



@api.get("/executive/whatsapp-stats")
async def whatsapp_stats(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    _: dict = Depends(require_role("koordinator")),
):
    base: dict = {}
    if from_date or to_date:
        rng: dict = {}
        if from_date:
            rng["$gte"] = from_date
        if to_date:
            rng["$lte"] = to_date + "T23:59:59"
        base["created_at"] = rng

    total = await db.wa_logs.count_documents(base)
    sent = await db.wa_logs.count_documents({**base, "status": True})
    failed = await db.wa_logs.count_documents({**base, "status": False, "skipped": {"$ne": True}})
    skipped = await db.wa_logs.count_documents({**base, "skipped": True})
    success_rate = round((sent / total) * 100, 1) if total else 0.0

    # last 24h / 7d / 30d
    now = datetime.now(timezone.utc)
    day1 = (now - timedelta(days=1)).isoformat()
    day7 = (now - timedelta(days=7)).isoformat()
    day30 = (now - timedelta(days=30)).isoformat()
    sent_24h = await db.wa_logs.count_documents({"created_at": {"$gte": day1}, "status": True})
    sent_7d = await db.wa_logs.count_documents({"created_at": {"$gte": day7}, "status": True})
    sent_30d = await db.wa_logs.count_documents({"created_at": {"$gte": day30}, "status": True})

    # by event type
    pipeline = [
        {"$match": base},
        {"$group": {"_id": {"event": "$event_type", "ok": "$status"}, "count": {"$sum": 1}}},
    ]
    bucket: dict = {}
    async for row in db.wa_logs.aggregate(pipeline):
        ev = row["_id"]["event"] or "unknown"
        ok = row["_id"]["ok"]
        rec = bucket.setdefault(ev, {"event": ev, "sent": 0, "failed": 0})
        if ok:
            rec["sent"] += row["count"]
        else:
            rec["failed"] += row["count"]
    by_event = sorted(bucket.values(), key=lambda r: r["sent"] + r["failed"], reverse=True)

    # last 10 failures
    failures = await db.wa_logs.find(
        {"status": False, "skipped": {"$ne": True}},
    ).sort("created_at", -1).limit(10).to_list(10)

    # current quota — try device endpoint, else latest log
    quota_remaining = None
    device = await fetch_device_info()
    try:
        if isinstance(device, dict) and isinstance(device.get("quota"), int):
            quota_remaining = device["quota"]
    except Exception:
        pass
    if quota_remaining is None:
        last_with_quota = await db.wa_logs.find(
            {"quota_remaining": {"$ne": None}},
        ).sort("created_at", -1).limit(1).to_list(1)
        if last_with_quota:
            quota_remaining = last_with_quota[0].get("quota_remaining")

    # daily timeline (last 14 days)
    timeline = []
    for i in range(13, -1, -1):
        start = (now.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=i)).isoformat()
        end = (now.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=i - 1)).isoformat()
        sc = await db.wa_logs.count_documents({"created_at": {"$gte": start, "$lt": end}, "status": True})
        fc = await db.wa_logs.count_documents({"created_at": {"$gte": start, "$lt": end}, "status": False, "skipped": {"$ne": True}})
        timeline.append({"date": start[:10], "sent": sc, "failed": fc})

    return {
        "total": total,
        "sent": sent,
        "failed": failed,
        "skipped": skipped,
        "success_rate": success_rate,
        "sent_24h": sent_24h,
        "sent_7d": sent_7d,
        "sent_30d": sent_30d,
        "by_event": by_event,
        "recent_failures": [doc_to_public(f) for f in failures],
        "quota_remaining": quota_remaining,
        "device": device if device else None,
        "timeline": timeline,
    }


@api.post("/admin/cleanup-test-tickets")
async def cleanup_test_tickets(user: dict = Depends(require_role("koordinator"))):
    """Delete tickets created by tests (judul/ticket_number starts with TEST_ or TEST-,
    or operator email contains 'test'). Cascades to activities, attachments, notifications."""
    test_filter = {
        "$or": [
            {"judul": {"$regex": "^TEST[_\\- ]", "$options": "i"}},
            {"judul": {"$regex": "playwright", "$options": "i"}},
            {"deskripsi": {"$regex": "^TEST[_\\- ]", "$options": "i"}},
        ]
    }
    to_delete = await db.tickets.find(test_filter, {"_id": 1, "ticket_number": 1}).to_list(5000)
    ids = [d["_id"] for d in to_delete]
    str_ids = [str(_id) for _id in ids]
    numbers = [d.get("ticket_number") for d in to_delete]

    if not ids:
        return {"deleted": 0, "tickets": []}

    await db.tickets.delete_many({"_id": {"$in": ids}})
    await db.activities.delete_many({"ticket_id": {"$in": str_ids}})
    await db.attachments.delete_many({"ticket_id": {"$in": str_ids}})
    await db.notifications.delete_many({"ticket_id": {"$in": str_ids}})

    await log_audit(
        user, "ticket", None, "cleanup_test",
        f"Cleanup {len(ids)} test ticket(s)",
        {"count": len(ids), "ticket_numbers": numbers[:50]},
    )
    return {"deleted": len(ids), "tickets": numbers}


# ---------- Dashboard ----------
@api.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(get_current_user)):
    base: dict = {}
    if user["role"] == "operator":
        base["operator_id"] = user["id"]

    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()

    total = await db.tickets.count_documents(base)
    today = await db.tickets.count_documents({**base, "submitted_at": {"$gte": today_start}})
    diproses = await db.tickets.count_documents({**base, "status": {"$in": ["Diajukan", "Diproses", "Menunggu Dokumen"]}})
    revisi = await db.tickets.count_documents({**base, "status": "Revisi"})
    selesai = await db.tickets.count_documents({**base, "status": {"$in": ["Selesai", "Disetujui"]}})
    ditolak = await db.tickets.count_documents({**base, "status": "Ditolak"})

    # SLA
    open_filter = {**base, "status": {"$nin": ["Selesai", "Disetujui", "Ditolak"]}}
    open_tickets = await db.tickets.find(open_filter, {"due_at": 1, "status": 1}).to_list(5000)
    on_time = almost = late = 0
    for t in open_tickets:
        s = sla_state(t.get("due_at"), t.get("status"))
        if s == "tepat_waktu":
            on_time += 1
        elif s == "hampir_terlambat":
            almost += 1
        elif s == "terlambat":
            late += 1

    # Monthly: last 6 months counts
    now = datetime.now(timezone.utc)
    months = []
    for i in range(5, -1, -1):
        y = now.year
        m = now.month - i
        while m <= 0:
            m += 12
            y -= 1
        start = datetime(y, m, 1, tzinfo=timezone.utc)
        if m == 12:
            end = datetime(y + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end = datetime(y, m + 1, 1, tzinfo=timezone.utc)
        count = await db.tickets.count_documents({
            **base,
            "submitted_at": {"$gte": start.isoformat(), "$lt": end.isoformat()},
        })
        months.append({"label": start.strftime("%b %Y"), "month": start.strftime("%Y-%m"), "count": count})

    # by status (for pie/legend)
    by_status = []
    for st in STATUSES:
        c = await db.tickets.count_documents({**base, "status": st})
        if c > 0:
            by_status.append({"status": st, "count": c})

    return {
        "total": total,
        "today": today,
        "diproses": diproses,
        "revisi": revisi,
        "selesai": selesai,
        "ditolak": ditolak,
        "sla": {"on_time": on_time, "almost": almost, "late": late},
        "monthly": months,
        "by_status": by_status,
    }


# ---------- Reports ----------
def _build_ticket_filter(
    from_date: Optional[str], to_date: Optional[str], status: Optional[str],
    layanan_id: Optional[str], kecamatan: Optional[str],
) -> dict:
    filt: dict = {}
    if from_date or to_date:
        rng = {}
        if from_date:
            rng["$gte"] = from_date
        if to_date:
            rng["$lte"] = to_date + "T23:59:59"
        filt["submitted_at"] = rng
    if status:
        filt["status"] = status
    if layanan_id:
        filt["layanan_id"] = layanan_id
    if kecamatan:
        filt["kecamatan"] = kecamatan
    return filt


@api.get("/reports/excel")
async def report_excel(
    from_date: Optional[str] = None, to_date: Optional[str] = None,
    status: Optional[str] = None, layanan_id: Optional[str] = None,
    kecamatan: Optional[str] = None,
    _: dict = Depends(require_role("koordinator")),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Pengajuan"
    headers = ["No. Ticket", "Tanggal", "Sekolah", "Kecamatan", "Operator", "Jenis Layanan", "Prioritas", "Status", "SLA (hari)", "Jatuh Tempo", "Selesai"]
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="09090B")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")

    filt = _build_ticket_filter(from_date, to_date, status, layanan_id, kecamatan)
    cursor = db.tickets.find(filt).sort("submitted_at", -1)
    async for t in cursor:
        ws.append([
            t.get("ticket_number"),
            t.get("submitted_at", "")[:10],
            t.get("sekolah_nama") or "-",
            t.get("kecamatan") or "-",
            t.get("operator_name") or "-",
            t.get("layanan_nama") or "-",
            t.get("prioritas"),
            t.get("status"),
            t.get("sla_days"),
            (t.get("due_at") or "")[:10],
            (t.get("closed_at") or "")[:10] if t.get("closed_at") else "-",
        ])
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"laporan-pengajuan-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api.get("/reports/pdf")
async def report_pdf(
    from_date: Optional[str] = None, to_date: Optional[str] = None,
    status: Optional[str] = None, layanan_id: Optional[str] = None,
    kecamatan: Optional[str] = None,
    _: dict = Depends(require_role("koordinator")),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=24, bottomMargin=24, leftMargin=24, rightMargin=24)
    styles = getSampleStyleSheet()
    title = Paragraph("<b>Laporan Pengajuan Layanan Dapodik</b>", styles["Title"])
    subtitle = Paragraph(
        f"Periode: {from_date or '-'} s.d. {to_date or '-'} &nbsp;&nbsp;|&nbsp;&nbsp; Dibuat: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        styles["Normal"],
    )

    rows = [["No. Ticket", "Tanggal", "Sekolah", "Operator", "Layanan", "Prioritas", "Status"]]
    filt = _build_ticket_filter(from_date, to_date, status, layanan_id, kecamatan)
    cursor = db.tickets.find(filt).sort("submitted_at", -1)
    async for t in cursor:
        rows.append([
            t.get("ticket_number"),
            (t.get("submitted_at") or "")[:10],
            (t.get("sekolah_nama") or "-")[:30],
            (t.get("operator_name") or "-")[:25],
            (t.get("layanan_nama") or "-")[:35],
            t.get("prioritas"),
            t.get("status"),
        ])
    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#09090B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E4E4E7")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    doc.build([title, Spacer(1, 6), subtitle, Spacer(1, 12), table])
    buf.seek(0)
    fname = f"laporan-pengajuan-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------- Health ----------
@api.get("/")
async def root():
    return {"status": "ok", "service": "Dapodik Ticketing API"}


# ---------- Seed ----------
async def seed():
    # indexes
    await db.users.create_index("email", unique=True)
    await db.tickets.create_index([("ticket_number", 1)], unique=True)
    await db.tickets.create_index([("submitted_at", -1)])
    await db.tickets.create_index([("operator_id", 1)])
    await db.activities.create_index([("ticket_id", 1), ("created_at", 1)])
    await db.notifications.create_index([("user_id", 1), ("created_at", -1)])
    await db.audit_logs.create_index([("created_at", -1)])
    await db.audit_logs.create_index([("entity", 1), ("created_at", -1)])
    await db.kb_articles.create_index([("updated_at", -1)])
    await db.kb_articles.create_index([("title", "text"), ("content", "text"), ("tags", "text")])
    await db.wa_logs.create_index([("created_at", -1)])
    await db.wa_logs.create_index([("status", 1), ("created_at", -1)])
    await db.wa_logs.create_index([("event_type", 1), ("created_at", -1)])

    # kecamatan
    for k in DEFAULT_KECAMATAN:
        if not await db.kecamatan.find_one({"nama": k}):
            await db.kecamatan.insert_one({"nama": k, "created_at": iso(now_utc())})

    # services
    for nama, sla, checklist, attachment_required in DEFAULT_SERVICES:
        existing = await db.services.find_one({"nama": nama})
        form_schema = DEFAULT_FORM_SCHEMAS.get(nama, [])
        if not existing:
            await db.services.insert_one({
                "nama": nama, "sla_days": sla, "deskripsi": None,
                "checklist": checklist, "form_schema": form_schema,
                "attachment_required": attachment_required,
                "created_at": iso(now_utc()),
            })
        else:
            updates: dict = {}
            if "checklist" not in existing:
                updates["checklist"] = checklist
            # normalize existing schema: NIP/NIK/NUPTK must be optional
            existing_schema = existing.get("form_schema") or []
            if not existing_schema:
                updates["form_schema"] = form_schema
            else:
                changed = False
                for f in existing_schema:
                    if isinstance(f, dict) and f.get("key") in OPTIONAL_FIELD_KEYS and f.get("required"):
                        f["required"] = False
                        changed = True
                if changed:
                    updates["form_schema"] = existing_schema
            # enforce attachment_required per default spec if not set or mismatched with defaults
            if "attachment_required" not in existing:
                updates["attachment_required"] = attachment_required
            if updates:
                await db.services.update_one({"_id": existing["_id"]}, {"$set": updates})

    # admin koordinator
    admin_email = os.environ["ADMIN_EMAIL"].lower()
    admin_pass = os.environ["ADMIN_PASSWORD"]
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "name": "Koordinator Tim",
            "email": admin_email,
            "password_hash": hash_password(admin_pass),
            "role": "koordinator",
            "active": True,
            "created_at": iso(now_utc()),
        })
    elif not verify_password(admin_pass, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_pass)}},
        )

    # sample sekolah + operators
    sample_sekolah = [
        {"nama": "SDN 01 Bogor Tengah", "npsn": "20220001", "kecamatan": "Bogor Tengah", "jenjang": "SD"},
        {"nama": "SMPN 02 Bogor Utara", "npsn": "20220002", "kecamatan": "Bogor Utara", "jenjang": "SMP"},
        {"nama": "SMAN 03 Bogor Selatan", "npsn": "20220003", "kecamatan": "Bogor Selatan", "jenjang": "SMA"},
    ]
    sekolah_ids = []
    for s in sample_sekolah:
        existing = await db.sekolah.find_one({"npsn": s["npsn"]})
        if not existing:
            res = await db.sekolah.insert_one({**s, "created_at": iso(now_utc())})
            sekolah_ids.append(str(res.inserted_id))
        else:
            sekolah_ids.append(str(existing["_id"]))

    sample_ops = [
        {"name": "Budi Santoso", "email": "operator1@dapodik.id", "password": "operator123", "sekolah_id": sekolah_ids[0]},
        {"name": "Siti Aminah", "email": "operator2@dapodik.id", "password": "operator123", "sekolah_id": sekolah_ids[1]},
        {"name": "Ahmad Rifai", "email": "operator3@dapodik.id", "password": "operator123", "sekolah_id": sekolah_ids[2]},
    ]
    for op in sample_ops:
        if not await db.users.find_one({"email": op["email"]}):
            await db.users.insert_one({
                "name": op["name"],
                "email": op["email"],
                "password_hash": hash_password(op["password"]),
                "role": "operator",
                "sekolah_id": op["sekolah_id"],
                "active": True,
                "created_at": iso(now_utc()),
            })

    logger.info("Seed complete")


@app.on_event("startup")
async def on_startup():
    await seed()


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


# Mount router and middleware
app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)
